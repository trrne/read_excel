from random import (
    randint,
    uniform
)

import openpyxl as pyxl
from openpyxl import (
    Workbook,
    worksheet,
)
from typing import (
    TypedDict,
    Any,
    TypeVar,
    Generic,
    overload
)

import tkinter as tk
from tkinter import (
    messagebox,
    filedialog
)

import asyncio


def choice(n: list) -> int:
    return randint(0, len(n)-1)


TSubject = TypeVar('TSubject')
TWeight = TypeVar('TWeight', int, float)


class IdeasListDict(TypedDict):
    nouns: list
    verbs: list


class Excel:
    def __init__(self, source_path: str) -> None:
        self.source_path: str = source_path
        self.file: Workbook
        self.sheet: worksheet
        self.sheets: list
        self.row: int
        self.column: int
        self.loop: int
        if self.source_path != None:
            self.file = pyxl.load_workbook(self.source_path)
            self.sheets = self.file.sheetnames
            if (name := self.sheets[0]) != None:
                self.sheet = self.file[name]
                self.row = self.sheet.max_row
                self.column = self.sheet.max_column

    def get_value(self, row: int, column: int) -> Any:
        return self.sheet.cell(row, column).value

    def read(self) -> IdeasListDict:
        (ns, vs) = ([], [])
        for i in range(3, self.row, 1):
            ns.append(self.get_value(i, 1))
            vs.append(self.get_value(i, 2))
        return IdeasListDict(
            nouns=[n for n in ns if n != None], verbs=[v for v in vs if v != None])

    # @staticmethod
    def generate_pair(self) -> None:
        rdata: IdeasListDict = self.read()
        (nouns, verbs) = (rdata['nouns'], rdata['verbs'])
        donestr = []
        try:
            for _ in range(self.loop):
                donestr.append(
                    f'{nouns[choice(nouns)]} + {verbs[choice(verbs)]}')
            output_label['text'] = str.join('\n', donestr)
        except:
            messagebox.showerror(
                title='loop count is empty.', message='please enter it.')


class MyApp(tk.Frame):
    def __init__(self, title: str, resolution: str) -> None:
        self.title: str = title
        self.resolution: str = resolution
        self.xlsx: str = None
        self.root = tk.Tk()
        if self.root != None:
            super().__init__(self.root)
            self.root.title(self.title)
            self.root.geometry(self.resolution)

    # @staticmethod
    def open_xlsx(self):
        self.xlsx = filedialog.askopenfilename(
            title='*.xlsxを開く',
            filetypes=[('XLSX Worksheet', '*.xlsx')],
            initialdir='./'
        )
        global open_btn, body
        open_btn.destroy()
        body = True

    def loop(self) -> None:
        self.root.mainloop()


async def body():
    print('run start')
    app = MyApp(title='アイデア出し', resolution='300x375')

    open_btn = tk.Button(app.root, text='*ファイルを開く', command=app.open_xlsx)
    xlsx: Excel = None

    if open_btn != None:
        open_btn.pack()

    if True:  # xlsx != None:
        if xlsx.source_path != None:
            name_label = tk.Label(
                app.root, text='選択しているファイル: ' + xlsx.source_path)
            name_label.pack()

        loop_input = tk.Entry(app.root, width=20)
        loop_input.insert(tk.END, 20)  # 入力できない
        loop_input.pack()

        xlsx.loop = int(loop_input.get())

        output_label = tk.Label()
        output_label.pack()
        run_btn = tk.Button(
            app.root, text='生成', command=app.xlsx.generate_pair)
        run_btn.pack()

    app.loop()

if __name__ == "__main__":
    body()
